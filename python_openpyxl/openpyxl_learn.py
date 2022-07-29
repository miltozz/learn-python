import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_low = {}


for row in range(2, product_list.max_row+1):
    product_number = int(product_list.cell(row, 1).value)
    product_quantity = int(product_list.cell(row, 2).value)
    product_value = product_list.cell(row, 3).value
    supplier_name = product_list.cell(row, 4).value

    # set. write to sheet. need whole object
    total_inv_value_pp = product_list.cell(row, 5)

    # how many products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # total product value per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += (
            product_quantity * product_value)
    else:
        total_value_per_supplier[supplier_name] = (
            product_quantity * product_value)

    # products with low inventory (<10)
    if product_quantity < 10:
        products_low[product_number] = product_quantity

    # calculate total value per product and write to new column
    total_inv_value_pp.value = product_quantity * product_value

print(products_per_supplier)
print(total_value_per_supplier)
print(products_low)

# write changes to file and save as.
# inv_file.save("inventory_updated.xlsx")
